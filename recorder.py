import torch
import matplotlib.pyplot as plt 
import openpyxl  
import os  

def save_model(model, file_name):
    '''
    Save the trained model to a file.

    Args:
        model (torch.nn.Module): The trained model to save.
        file_name (str): The file name to save the model to.

    Returns:
        None
    '''
    torch.save(model, file_name)  # Save the model to the specified file

def save_graph(x_data, y_data, file_name):
    '''
    Save the graph of rewards over epochs.

    Args:
        x_data (list): The x-axis data (epochs).
        y_data (list): The y-axis data (rewards).
        file_name (str): The file name to save the graph to.

    Returns:
        None
    '''
    plt.xlabel("epoch")  # Set the label for the x-axis
    plt.ylabel("reward")  # Set the label for the y-axis
    plt.plot(x_data, y_data)  # Plot the data
    plt.savefig(file_name)  # Save the plot to the specified file
    plt.show()  # Display the plot

def save_data(record_rewards, file_name):
    '''
    Save the reward data to an Excel file.

    Args:
        record_rewards (list): List of rewards to save.
        file_name (str): The file name to save the data to.

    Returns:
        None
    '''
    is_file = os.path.isfile(file_name)  # Check if the file already exists
    if not is_file:
        wb = openpyxl.Workbook()  # Create a new workbook
        wb.save(file_name)  # Save the workbook with the specified file name

    wb = openpyxl.load_workbook(file_name)  # Load the workbook
    ws_rewards = wb.create_sheet(index=0, title="average rewards")  # Create a new sheet for the rewards
    ws_rewards.cell(1, 1).value = "average rewards"  # Set the header for the rewards column
    for i, reward in enumerate(record_rewards, start=2):  # Write the rewards to the sheet
        ws_rewards.cell(i, 1).value = reward
    wb.save(file_name)  # Save the workbook
